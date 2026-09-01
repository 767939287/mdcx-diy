"""ASIN 数据库查询缓存测试：内存索引只建一次 + mtime 失效自动重建。"""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.core.amazon_database import invalidate_asin_cache, query_asin_database

pytestmark = pytest.mark.asyncio


def _make_xlsx(path: Path, rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(("number", "asin", "product_url", "title", "poster_url", "search_keyword"))
    for row in rows:
        ws.append(row)
    wb.save(path)


async def test_query_hits_and_index_built_once(monkeypatch, tmp_path):
    """两次查询只解析一次文件（第二次命中内存索引）"""
    db_path = tmp_path / "asin.xlsx"
    _make_xlsx(db_path, [("ABC-123", "B001", "u1", "t1", "p1", "k1"), ("DEF-456", "B002", "u2", "t2", "p2", "k2")])

    load_calls = []
    import openpyxl

    real_load = openpyxl.load_workbook

    def counting_load(*args, **kwargs):
        load_calls.append(1)
        return real_load(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", counting_load)
    invalidate_asin_cache(db_path)

    r1 = await query_asin_database(number="ABC-123", excel_path=db_path)
    r2 = await query_asin_database(number="ABC-123", excel_path=db_path)
    r3 = await query_asin_database(asin="B002", excel_path=db_path)

    assert len(r1) == 1 and r1[0]["asin"] == "B001"
    assert r2 == r1
    assert len(r3) == 1 and r3[0]["number"] == "DEF-456"
    assert len(load_calls) == 1  # 两次 number 查询 + 一次 asin 查询只 load 一次


async def test_cache_invalidates_on_file_change(monkeypatch, tmp_path):
    """文件更新（mtime 变化）后自动重建索引，查到新数据"""
    db_path = tmp_path / "asin.xlsx"
    _make_xlsx(db_path, [("ABC-123", "B001", "u1", "t1", "p1", "k1")])
    invalidate_asin_cache(db_path)

    r1 = await query_asin_database(number="ABC-123", excel_path=db_path)
    assert len(r1) == 1 and r1[0]["title"] == "t1"

    # 追加一行（mtime 变化）
    wb = load_workbook(db_path)
    ws = wb.active
    ws.append(("GHI-789", "B003", "u3", "t3", "p3", "k3"))
    wb.save(db_path)

    r2 = await query_asin_database(number="GHI-789", excel_path=db_path)
    assert len(r2) == 1 and r2[0]["asin"] == "B003"
    # 旧行还在
    r3 = await query_asin_database(number="ABC-123", excel_path=db_path)
    assert len(r3) == 1


async def test_case_insensitive_lookup(monkeypatch, tmp_path):
    """番号/ASIN 大小写不敏感查询"""
    db_path = tmp_path / "asin.xlsx"
    _make_xlsx(db_path, [("abc-001", "b1xx", "u", "t", "p", "k")])
    invalidate_asin_cache(db_path)

    assert len(await query_asin_database(number="ABC-001", excel_path=db_path)) == 1
    assert len(await query_asin_database(asin="B1XX", excel_path=db_path)) == 1


async def test_missing_file_returns_empty(monkeypatch, tmp_path):
    """文件不存在返回空且不炸"""
    invalidate_asin_cache(tmp_path / "nonexistent.xlsx")
    assert await query_asin_database(number="X-1", excel_path=tmp_path / "nonexistent.xlsx") == []


def test_invalidate_is_safe_without_prior_cache(tmp_path):
    """无缓存时 invalidate 幂等"""
    invalidate_asin_cache(tmp_path / "whatever.xlsx")
    invalidate_asin_cache(tmp_path / "whatever.xlsx")
