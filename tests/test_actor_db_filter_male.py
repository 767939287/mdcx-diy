import asyncio
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mdcx.config.resources import COL_JP, COL_TMDBID, DB_HEADERS
from mdcx.core import tmdb_actor
from mdcx.tools import actor_db_tool

_XML = """<?xml version="1.0" encoding="UTF-8"?>
<actor-mapping>
  <actor>
    <a zh_cn="阿部純子" zh_tw="阿部純子" jp="阿部純子" keyword="阿部純子" tmdb_id="1417328" verified="1" />
    <a zh_cn="阿部涼音" zh_tw="阿部涼音" jp="阿部涼音" keyword="阿部涼音" tmdb_id="1417329" verified="1" />
  </actor>
</actor-mapping>
"""


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


@pytest.fixture(autouse=True)
def _reset_actor_db_row_index():
    with tmdb_actor._ACTOR_DB_ROW_INDEX_LOCK:
        tmdb_actor._ACTOR_DB_ROW_INDEX.clear()


@pytest.fixture
def _avdb_xml(tmp_path: Path):
    path = tmp_path / "mapping.xml"
    path.write_text(_XML, encoding="utf-8")
    return path


def _write_db(path: Path, rows, headers=DB_HEADERS):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def _read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return rows


def _mock_tmdb(
    monkeypatch,
    genders: dict[int, int | None],
    identity_names: dict[int, str] | None = None,
    identity_spy: list[int] | None = None,
):
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("http://base", "test-key"))
    calls: list[int] = []

    async def fake_fetch(pid, base_url, api_key, client):
        calls.append(pid)
        return genders.get(pid)

    async def fake_fetch_identity(pid, base_url, api_key, client):
        if identity_spy is not None:
            identity_spy.append(pid)
        if identity_names is None:
            return None  # 与旧行为一致：身份不可知 -> 放行
        name = identity_names.get(pid)
        if name is None:
            return None
        return {"gender": None, "name": name, "original_name": name, "also_known_as": []}

    monkeypatch.setattr(actor_db_tool, "fetch_person_gender", fake_fetch)
    monkeypatch.setattr(actor_db_tool, "fetch_person_identity", fake_fetch_identity)
    return calls


def test_sync_skips_male_when_filter_male(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert result.created == 1
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 1
    assert rows[0][COL_JP] == "阿部涼音"


def test_sync_keeps_female_and_unknown(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: 1, 1417329: 0})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_keeps_when_gender_unknown(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(monkeypatch, {1417328: None, 1417329: None})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_no_filter_when_no_tmdb_key(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 0
    assert result.created == 2


def test_sync_does_not_requery_existing_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["阿部純子", "阿部純子", "", "", "", 1417328, "", "", ""]])
    calls = _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert 1417328 not in calls  # 本地已有该 tmdbid，不重复请求性别
    assert result.skipped_male == 0
    assert result.created == 1


def test_clean_removes_male_and_backs_up(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [["男优A", "男优A", "", "", "", 1001, "", "", ""], ["女优B", "女优B", "", "", "", 1002, "", "", ""]],
    )
    _mock_tmdb(monkeypatch, {1001: 2, 1002: 1})
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优B"]
    wb = load_workbook(_tmp_actor_db)
    assert "男优备份" in wb.sheetnames
    backup = list(wb["男优备份"].iter_rows(min_row=1, values_only=True))
    wb.close()
    assert len(backup) == 1
    assert backup[0][0] == "男优A"


def test_clean_keeps_female_unknown_and_missing(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["女优A", "女优A", "", "", "", 2001, "", "", ""],
            ["未知B", "未知B", "", "", "", 2002, "", "", ""],
            ["无idC", "无idC", "", "", "", "", "", "", ""],
        ],
    )
    _mock_tmdb(monkeypatch, {2001: 1, 2002: None})
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 3


def test_clean_no_tmdb_key_noop(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["男优A", "男优A", "", "", "", 3001, "", "", ""]])
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    assert len(_read_rows(_tmp_actor_db)) == 1


def test_clean_male_list_works_without_tmdb_key(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["吉村卓", "吉村卓", "", "", "", "", "", "", ""],
            ["女优A", "女优A", "", "", "", "", "", "", ""],
        ],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("", ""))
    _mock_male_list(monkeypatch, ["吉村卓"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优A"]


def test_clean_limit_applies(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [["男优A", "男优A", "", "", "", 4001, "", "", ""], ["男优B", "男优B", "", "", "", 4002, "", "", ""]],
    )
    _mock_tmdb(monkeypatch, {4001: 2, 4002: 2})
    result = asyncio.run(actor_db_tool.clean_male_actors(limit=1))
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert len(rows) == 1


def test_clean_idempotent(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["女优A", "女优A", "", "", "", 5001, "", "", ""]])
    _mock_tmdb(monkeypatch, {5001: 1})
    asyncio.run(actor_db_tool.clean_male_actors())
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 0
    assert len(_read_rows(_tmp_actor_db)) == 1


def _mock_male_list(monkeypatch, names: list[str]):
    monkeypatch.setattr(actor_db_tool, "_load_male_actor_set", lambda: {n.casefold() for n in names})


def test_clean_removes_male_by_list_without_tmdbid(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["吉村卓", "吉村卓", "", "", "", "", "", "", ""],
            ["女优A", "女优A", "", "", "", "", "", "", ""],
        ],
    )
    _mock_tmdb(monkeypatch, {})
    _mock_male_list(monkeypatch, ["吉村卓"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["女优A"]
    wb = load_workbook(_tmp_actor_db)
    backup = list(wb["男优备份"].iter_rows(min_row=1, values_only=True))
    wb.close()
    assert [b[0] for b in backup] == ["吉村卓"]


def test_clean_male_list_beats_tmdb_gender0(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["加藤鷹", "加藤鷹", "", "", "", 6001, "", "", ""]])
    _mock_tmdb(monkeypatch, {6001: 0})  # TMDB 未标性别，但名单命中
    _mock_male_list(monkeypatch, ["加藤鷹"])
    result = asyncio.run(actor_db_tool.clean_male_actors())
    assert result.removed_male == 1
    assert len(_read_rows(_tmp_actor_db)) == 0


def test_clean_male_list_only_name_checked(_tmp_actor_db: Path, monkeypatch):
    _write_db(_tmp_actor_db, [["吉村卓", "吉村卓", "", "", "", 7001, "", "", ""]])
    calls = _mock_tmdb(monkeypatch, {7001: 2})
    _mock_male_list(monkeypatch, ["吉村卓"])
    asyncio.run(actor_db_tool.clean_male_actors())
    assert 7001 not in calls  # 名单命中后不应再请求 TMDB


def test_sync_skips_male_by_list_without_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_male_list(monkeypatch, ["阿部純子"])
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert result.created == 1
    rows = _read_rows(_tmp_actor_db)
    assert [r[COL_JP] for r in rows] == ["阿部涼音"]


def test_sync_male_list_avoids_tmdb_request(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_male_list(monkeypatch, ["阿部純子"])
    calls = _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1})
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_male == 1
    assert 1417328 not in calls  # 名单命中，不应发起 TMDB gender 请求


class _FakeResp:
    def __init__(self, status: int):
        self.status = status

    async def __aenter__(self):
        return self

    async def __aexit__(self, *args):
        return None

    def release(self):
        return None


def _mock_verify_network(monkeypatch, status_map: dict[int, int]):
    """mock verify_tmdb_ids 的 TMDB person/{id} 请求。status_map: tmdbid -> http status。"""
    import aiohttp

    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    class _FakeGet:
        def __call__(self, url, timeout=None):
            import re

            m = re.search(r"/person/(\d+)", url)
            pid = int(m.group(1)) if m else 0
            return _FakeResp(status_map.get(pid, 200))

    class _FakeClient:
        def __init__(self):
            self.get = _FakeGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)


def test_verify_tmdbid_clears_invalid_ids(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["桃乃木香奈", "", "", "", "", 2616715, "https://www.themoviedb.org/person/2616715", "", ""],
            ["三佳詩", "", "", "", "", 6231965, "https://www.themoviedb.org/person/6231965", "", ""],  # 404 失效
            ["涼森れむ", "", "", "", "", 2640963, "https://www.themoviedb.org/person/2640963", "", ""],
        ],
    )
    _mock_verify_network(monkeypatch, {2616715: 200, 6231965: 404, 2640963: 200})
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.checked == 3
    assert result.invalid == 1
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[0]: r[5] for r in rows}
    assert by_jp["桃乃木香奈"] == 2616715  # 有效保留
    assert by_jp["三佳詩"] is None  # 失效清除
    assert by_jp["涼森れむ"] == 2640963  # 有效保留


def test_verify_tmdbid_keeps_all_valid(_tmp_actor_db: Path, monkeypatch):
    _write_db(
        _tmp_actor_db,
        [
            ["桃乃木香奈", "", "", "", "", 2616715, "https://www.themoviedb.org/person/2616715", "", ""],
            ["涼森れむ", "", "", "", "", 2640963, "https://www.themoviedb.org/person/2640963", "", ""],
        ],
    )
    _mock_verify_network(monkeypatch, {2616715: 200, 2640963: 200})
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.invalid == 0
    rows = _read_rows(_tmp_actor_db)
    assert {r[5] for r in rows} == {2616715, 2640963}


def test_verify_tmdbid_network_error_keeps_id(_tmp_actor_db: Path, monkeypatch):
    """网络失败(非404)保守保留 id，不误清。"""
    import aiohttp

    _write_db(
        _tmp_actor_db,
        [["某演员", "", "", "", "", 1001, "https://www.themoviedb.org/person/1001", "", ""]],
    )
    monkeypatch.setattr(actor_db_tool, "_resolve_tmdb_config", lambda: ("https://api.tmdb.org", "test-key"))

    class _BoomGet:
        def __call__(self, url, timeout=None):
            raise aiohttp.ClientError("network down")

    class _FakeClient:
        def __init__(self):
            self.get = _BoomGet()

        async def __aenter__(self):
            return self

        async def __aexit__(self, *args):
            return None

    monkeypatch.setattr(aiohttp, "ClientSession", _FakeClient)
    result = asyncio.run(actor_db_tool.verify_tmdb_ids())
    assert result.invalid == 0
    rows = _read_rows(_tmp_actor_db)
    assert rows[0][5] == 1001  # 保留


def test_sync_keeps_tmdbid_when_identity_matches(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    calls = _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "阿部純子", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] == 1417328
    assert by_jp["阿部涼音"] == 1417329
    assert 1417328 in calls and 1417329 in calls


def test_sync_drops_tmdbid_when_identity_mismatch(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Christa Allen", 1417329: "阿部涼音"},  # 错误映射样本：平山加奈->Christa Allen
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 1
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] in (None, "")  # 身份不匹配，丢弃该 id
    assert by_jp["阿部涼音"] == 1417329  # 匹配的保留


def test_sync_verify_disabled_keeps_tmdbid(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Christa Allen", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml), verify_tmdbid=False))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["阿部純子"] == 1417328  # 关闭校验时原样写入


def test_sync_identity_match_with_katakana_roman(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    """片假名音译 ↔ 英文原名视为同人，不丢弃 id。"""
    xml = (
        _XML.replace('tmdb_id="1417328"', 'tmdb_id="1417328"')
        .replace('jp="阿部純子"', 'jp="キャシー・ヘブン"')
        .replace('zh_cn="阿部純子"', 'zh_cn="キャシー・ヘブン"')
    )
    xml_path = _avdb_xml.parent / "mapping2.xml"
    xml_path.write_text(xml, encoding="utf-8")
    _mock_tmdb(
        monkeypatch,
        {1417328: 1, 1417329: 1},
        identity_names={1417328: "Cathy Heaven", 1417329: "阿部涼音"},
    )
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(xml_path)))
    assert result.skipped_tmdbid == 0
    rows = _read_rows(_tmp_actor_db)
    by_jp = {r[COL_JP]: r[COL_TMDBID] for r in rows}
    assert by_jp["キャシー・ヘブン"] == 1417328  # 片假名↔英文放行


def test_sync_existing_tmdbid_not_requeried_for_identity(_tmp_actor_db: Path, _avdb_xml: Path, monkeypatch):
    """本地已存在的 tmdbid 不重复做身份反查。"""
    _write_db(_tmp_actor_db, [["阿部純子", "阿部純子", "", "", "", 1417328, "", "", ""]])
    spy: list[int] = []
    _mock_tmdb(monkeypatch, {1417328: 2, 1417329: 1}, identity_spy=spy)
    result = asyncio.run(actor_db_tool.sync_from_avdb("file", str(_avdb_xml)))
    assert result.skipped_tmdbid == 0
    assert 1417328 not in spy  # 已在库中的 id 不重复反查
    assert 1417329 in spy  # 新 id 需要校验
