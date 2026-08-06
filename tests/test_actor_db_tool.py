import asyncio
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mdcx.core import tmdb_actor
from mdcx.models.flags import Flags
from mdcx.signals import signal
from mdcx.tools import actor_db_tool


@pytest.fixture
def _reset_stop_flags():
    Flags.stop_requested = False
    signal.stop = False
    yield
    Flags.stop_requested = False
    signal.stop = False


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


def test_collect_actors_from_nfo_dir_gathers_and_dedups(tmp_path: Path):
    sub = tmp_path / "sub"
    sub.mkdir(parents=True, exist_ok=True)
    (tmp_path / "a.nfo").write_text(
        '<?xml version="1.0"?><movie><actor><name>三上悠亚</name></actor><actor><name>明日花绮罗</name></actor></movie>',
        encoding="utf-8",
    )
    (sub / "b.nfo").write_text(
        '<?xml version="1.0"?><movie><actor><name>三上悠亚</name></actor><actor><name>桥本有菜</name></actor></movie>',
        encoding="utf-8",
    )
    (tmp_path / "not.nfo").write_text("hello", encoding="utf-8")

    import asyncio

    actors = asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path))

    assert set(actors) == {"三上悠亚", "明日花绮罗", "桥本有菜"}


def test_collect_actors_from_nfo_dir_empty_dir(tmp_path: Path):
    import asyncio

    assert asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path)) == []


def test_collect_actors_from_nfo_dir_non_existent(tmp_path: Path):
    import asyncio

    assert asyncio.run(actor_db_tool.collect_actors_from_nfo_dir(tmp_path / "missing")) == []


@pytest.mark.asyncio
async def test_run_with_empty_names(_tmp_actor_db: Path):
    result = await actor_db_tool.run([])
    assert result.total == 0
    assert result.translated == 0
    assert result.linked == 0
    assert result.skipped == 0


@pytest.mark.asyncio
async def test_run_skips_actor_without_tmdbid(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    monkeypatch.setattr(tmdb_actor.resources, "actor_db_reverse_index", None)

    result = await actor_db_tool.run(["未知演员"], translate=True, link=True)

    assert result.total == 1
    assert result.skipped == 1
    assert result.translated == 0
    assert result.linked == 0


@pytest.mark.asyncio
async def test_run_translate_backfills_zh_names(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_translations(pid, base_url, api_key, client):
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _fake_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _no_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=True, link=False)

    assert result.total == 1
    assert result.translated == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[1] == "三上悠亚"
            assert row[2] == "三上悠亞"
            break
    wb.close()


@pytest.mark.asyncio
async def test_run_link_backfills_href(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="三上悠亚", zh_tw="三上悠亞", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {
            "zh_cn": "三上悠亚",
            "zh_tw": "三上悠亞",
            "keyword": "",
            "href": "",
            "tmdbid": 12345,
            "tmdb_url": "",
        }
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_link(actor_name: str) -> str:
        return "https://www.libredmm.com/actresses/mikami-yua"

    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _fake_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=False, link=True)

    assert result.total == 1
    assert result.linked == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[4] == "https://www.libredmm.com/actresses/mikami-yua"
            break
    wb.close()


@pytest.mark.asyncio
async def test_run_skips_when_translate_disabled(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    call_count = {"translations": 0}

    async def _counting_translations(pid, base_url, api_key, client):
        call_count["translations"] += 1
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _counting_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _no_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=False, link=False)

    assert result.total == 1
    assert result.skipped == 1
    assert call_count["translations"] == 0


@pytest.mark.asyncio
async def test_run_translate_and_link_together(monkeypatch: pytest.MonkeyPatch, _tmp_actor_db: Path):
    await tmdb_actor.update_actor_db_row(jp="三上悠亜", zh_cn="", zh_tw="", href="", tmdbid=12345)
    tmdb_actor.resources.actor_db = {
        "三上悠亜": {"zh_cn": "", "zh_tw": "", "keyword": "", "href": "", "tmdbid": 12345, "tmdb_url": ""}
    }
    tmdb_actor.resources.actor_db_reverse_index = None

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    async def _fake_translations(pid, base_url, api_key, client):
        return {"zh_cn": "三上悠亚", "zh_tw": "三上悠亞"}

    async def _fake_link(actor_name: str) -> str:
        return "https://www.libredmm.com/actresses/mikami-yua"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", _fake_translations)
    monkeypatch.setattr(actor_db_tool, "fetch_libredmm_link", _fake_link)

    result = await actor_db_tool.run(["三上悠亜"], translate=True, link=True)

    assert result.total == 1
    assert result.translated == 1
    assert result.linked == 1

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "").strip() == "三上悠亜":
            assert row[1] == "三上悠亚", f"zh_cn actual: {row[1]}"
            assert row[2] == "三上悠亞", f"zh_tw actual: {row[2]}"
            assert row[4] == "https://www.libredmm.com/actresses/mikami-yua", f"href actual: {row[4]}"
            break
    wb.close()


async def _no_link(actor_name: str) -> str:
    return ""


@pytest.mark.asyncio
async def test_update_actor_db_row_skips_placeholder_name(_tmp_actor_db: Path):
    """占位符名字被语义清洗为空时应跳过写入。"""
    status = await tmdb_actor.update_actor_db_row(jp="素人奥様", zh_cn="", zh_tw="", tmdbid=None)
    assert status == "skipped_placeholder"
    assert not _tmp_actor_db.exists()  # 未创建数据库文件，即未写入任何行


@pytest.mark.asyncio
async def test_update_actor_db_row_cleans_series_tag(_tmp_actor_db: Path):
    """写入时剥离名字中的系列标签。"""
    status = await tmdb_actor.update_actor_db_row(jp="本田仁美(パコパコママ)", zh_cn="", zh_tw="", tmdbid=None)
    assert status in ("inserted_new_row", "updated_zh_cn")
    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert rows and rows[0][0] == "本田仁美"


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_stop_cancels_pending_and_saves(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """run_actor_db_xlsx 在手动停止后取消 pending 并保存已处理部分。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    for jp, pid in [("演员甲", 11), ("演员乙", 12), ("演员丙", 13), ("演员丁", 14), ("演员戊", 15)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    calls: list[str] = []
    write_lock = asyncio.Lock()

    async def fake_translations(pid, base_url, api_key, client):
        async with write_lock:
            calls.append(str(pid))
            if len(calls) >= 3:
                Flags.stop_requested = True
        await asyncio.sleep(0.01)
        return {"zh_cn": f"中文{pid}", "zh_tw": f"中文{pid}"}

    async def fake_person_url(*args, **kwargs):
        return "https://www.themoviedb.org/person/1"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", fake_translations)
    monkeypatch.setattr(actor_db_tool, "_tmdb_person_url", fake_person_url)

    await actor_db_tool.run_actor_db_xlsx("translate")

    assert Flags.stop_requested is True
    assert len(calls) <= 5  # 停止后不再提交新的

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    saved = [r for r in rows if str(r[1] or "").strip()]
    assert len(saved) >= 2  # 已处理部分被保存
    for r in saved:
        assert str(r[1]).startswith("中文")


@pytest.mark.asyncio
async def test_run_actor_db_xlsx_limit_slices_and_reruns_idempotently(
    _tmp_actor_db: Path, monkeypatch: pytest.MonkeyPatch, _reset_stop_flags
):
    """limit 限量分片：仅处理前 limit 条，重跑时不重复已处理条目。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "演员数据库"
    ws.append(["日文原名", "中文名", "繁体名", "别名", "链接", "tmdbid", "tmdb url", "出生日期", "简介"])
    for jp, pid in [("演员甲", 11), ("演员乙", 12), ("演员丙", 13), ("演员丁", 14)]:
        ws.append([jp, "", "", "", "", str(pid), "", "", ""])
    wb.save(_tmp_actor_db)
    wb.close()

    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_key", "fake-key")
    monkeypatch.setattr(tmdb_actor.manager.config, "tmdb_api_base", "api.tmdb.org")

    processed: list[str] = []

    async def fake_translations(pid, base_url, api_key, client):
        processed.append(str(pid))
        await asyncio.sleep(0)
        return {"zh_cn": f"中文{pid}", "zh_tw": f"中文{pid}"}

    async def fake_person_url(*args, **kwargs):
        return "https://www.themoviedb.org/person/1"

    monkeypatch.setattr(actor_db_tool, "_fetch_person_translations", fake_translations)
    monkeypatch.setattr(actor_db_tool, "_tmdb_person_url", fake_person_url)

    # 第一轮：限量 2 条
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 2

    # 第二轮：重跑，已处理的条目不再进入，仅处理剩余
    processed.clear()
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 2

    # 第三轮：全部处理完，无剩余
    processed.clear()
    await actor_db_tool.run_actor_db_xlsx("translate", limit=2)
    assert len(processed) == 0

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert all(str(r[1] or "").startswith("中文") for r in rows)
