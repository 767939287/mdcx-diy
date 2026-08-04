import asyncio
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mdcx.config.resources import COL_BIO, COL_BIRTH_DATE, DB_HEADERS
from mdcx.core import tmdb_actor
from mdcx.utils.xml_avdb import (
    clean_actor_value,
    extract_birth_date,
    parse_avdb_actor_mapping,
    strip_age_and_birth,
)


@pytest.fixture
def _tmp_actor_db(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setattr(tmdb_actor.manager, "data_folder", tmp_path)
    monkeypatch.setattr(tmdb_actor.resources, "actor_db", {})
    userdata = tmp_path / "userdata"
    userdata.mkdir(parents=True, exist_ok=True)
    return userdata / "actor_database.xlsx"


@pytest.fixture(autouse=True)
def _reset_actor_db_row_index():
    with tmdb_actor._ACTOR_DB_ROW_INDEX_LOCK:
        tmdb_actor._ACTOR_DB_ROW_INDEX.clear()


def test_db_headers_has_nine_columns():
    assert len(DB_HEADERS) == 9
    assert DB_HEADERS[7] == "出生日期"
    assert DB_HEADERS[8] == "简介"
    assert COL_BIRTH_DATE == 7
    assert COL_BIO == 8


def test_update_actor_db_row_writes_birth_date_and_bio(_tmp_actor_db: Path):
    status = asyncio.run(
        tmdb_actor.update_actor_db_row(
            jp="新演员",
            zh_cn="新演员",
            tmdbid=99999,
            birth_date="1990-01-01",
            bio="身高160cm",
        )
    )
    assert status == "inserted_new_row"

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=COL_BIRTH_DATE + 1).value == "1990-01-01"
    assert ws.cell(row=2, column=COL_BIO + 1).value == "身高160cm"
    wb.close()


def test_update_actor_db_row_keeps_existing_birth_date(_tmp_actor_db: Path):
    asyncio.run(tmdb_actor.update_actor_db_row(jp="演员A", birth_date="1990-01-01", bio="旧简介"))
    asyncio.run(tmdb_actor.update_actor_db_row(jp="演员A", birth_date="2000-02-02", bio="新简介"))

    wb = load_workbook(_tmp_actor_db)
    ws = wb.active
    assert ws.cell(row=2, column=COL_BIRTH_DATE + 1).value == "1990-01-01"
    assert ws.cell(row=2, column=COL_BIO + 1).value == "旧简介"
    wb.close()


_SAMPLE_XML = """<?xml version="1.0" encoding="UTF-8"?>
<actor-mapping>
  <actor>
    <a zh_cn="阿部純子" zh_tw="阿部純子" jp="阿部純子" keyword="Abe Junko,阿部純子" tmdb_id="1417328" verified="1" bio_graphy="安部純子（あべじゅんこ / Abe Junko），1993年06月05日出生，33岁，身高158cm，三围B86/W78/H82，籍贯东京都。" />
    <a zh_cn="阿部涼音" zh_tw="阿部涼音" jp="阿部涼音" keyword="阿部涼音" bio_graphy="阿部涼音（あべすずね），身高157cm，三围B88/W58/H87。" />
    <a jp="无名字段" />
  </actor>
  <actor-blacklist>
    黑名单演员
  </actor-blacklist>
</actor-mapping>
"""


def test_parse_count_and_blacklist_ignored():
    actors = parse_avdb_actor_mapping(_SAMPLE_XML)
    assert len(actors) == 3
    assert actors[0].tmdb_id == "1417328"
    assert actors[1].bio_graphy


def test_parse_missing_fields_are_empty():
    actors = parse_avdb_actor_mapping(_SAMPLE_XML)
    missing = actors[2]
    assert missing.zh_cn == ""
    assert missing.keyword == ""
    assert missing.tmdb_id == ""


def test_parse_invalid_xml_raises_value_error():
    try:
        parse_avdb_actor_mapping("<actor-mapping><actor>")
    except ValueError:
        return
    raise AssertionError("expected ValueError for invalid xml")


def test_extract_birth_date_full_formats():
    assert extract_birth_date("1993年06月05日出生") == "1993-06-05"
    assert extract_birth_date("1993年6月5日出生") == "1993-06-05"
    assert extract_birth_date("1993.10.18 出生") == "1993-10-18"
    assert extract_birth_date("1993/1/5出生") == "1993-01-05"
    assert extract_birth_date("1993-06-05 出生") == "1993-06-05"


def test_extract_birth_date_partial_and_missing():
    assert extract_birth_date("1993年6月出生") == "1993-06"
    assert extract_birth_date("出生于1993年") == "1993"
    assert extract_birth_date("身高158cm，三围B86/W78/H82") == ""
    assert extract_birth_date("") == ""


def test_strip_age_and_birth_removes_dynamic_parts():
    bio = "安部純子，1993年06月05日出生，33岁，身高158cm，三围B86/W78/H82，籍贯东京都。"
    cleaned = strip_age_and_birth(bio, "1993-06-05")
    assert "1993" not in cleaned
    assert "33岁" not in cleaned
    assert "身高158cm" in cleaned
    assert "三围B86/W78/H82" in cleaned


def test_strip_age_and_birth_keeps_other_age_digits():
    bio = "身高158cm，三围B86/W78/H82。"
    cleaned = strip_age_and_birth(bio)
    assert "158cm" in cleaned
    assert "86/W78" in cleaned


def test_clean_actor_value_decodes_double_entities():
    assert clean_actor_value("&amp;quot;美咲&amp;quot;") == '"美咲"'


def test_clean_actor_value_removes_control_and_backslash_escapes():
    assert clean_actor_value("  美咲\n\t\x00\x1f  ") == "美咲"
    assert clean_actor_value("\\u4f50\\x41\\n山田") == "山田"


def test_clean_actor_value_trim():
    assert clean_actor_value("  三上悠亚  ") == "三上悠亚"
    assert clean_actor_value("") == ""
