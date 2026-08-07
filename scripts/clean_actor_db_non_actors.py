"""从出厂演员库删除「非演员」混入行。

第 1 类：导演/幕后人员——TMDB known_for_department != Acting 的 AV 从业者。
第 2 类：描述词/占位符——非人名的标签词（本妻/多数/女優不明 等）。
第 3 类：确证的非 AV 主流人物（埃里克·坎通纳/艾琳娜·拉尼娜 等）。

删除前在 /tmp 保留原始库副本（工作区安全网，不写入仓库）。
用法:
    python scripts/clean_actor_db_non_actors.py            # 预览
    python scripts/clean_actor_db_non_actors.py --apply    # 执行删除
    python scripts/clean_actor_db_non_actors.py --db <path> --apply  # 测试用副本
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

from mdcx.config.resources import get_actor_db_sheet  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"
NON_ACTING_FILE = Path("/tmp/non_acting.txt")

# 第 2 类：描述词/占位符（人工甄别，非人名）
_DESC_ROWS = {
    "愛嬌抜群のおんな",
    "愛人2",
    "坂道系のツンデレ女子大生",
    "本妻",
    "多数",
    "高学歴娘",
    "宮崎 後藤夫妻 デビ みずき はるか りえ",
    "金持ちを食い荒らす女",
    "可愛すぎる美女",
    "女優不明",
    "女優多数",
    "人物不明",
    "貧乳美女",
    "神待ち娘",
    "円光娘",
    "S級インテリ美女",
}

# 第 3 类：确证的非 AV 主流人物（有 id，人工确认后硬删）
_NON_AV_NAMES = {
    "埃里克·坎通纳",  # 足球明星
    "艾琳娜·拉尼娜",  # 主流演员
}


def collect_non_acting_ids() -> set[int]:
    """读第 1 类清单，返回待删 tmdbid。"""
    ids: set[int] = set()
    if not NON_ACTING_FILE.exists():
        return ids
    for line in NON_ACTING_FILE.read_text(encoding="utf-8").splitlines():
        parts = line.rstrip().split("\t")
        if len(parts) == 3 and parts[1].isdigit():
            ids.add(int(parts[1]))
    return ids


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="删除出厂库中的非演员混入行")
    parser.add_argument("--apply", action="store_true", help="执行删除")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="目标库路径（默认出厂库）")
    args = parser.parse_args(argv)

    db_path = args.db
    if not db_path.exists():
        print(f"数据库不存在: {db_path}")
        return 1

    non_acting_ids = collect_non_acting_ids()
    if not non_acting_ids:
        print(f"⚠️ 第 1 类清单为空: {NON_ACTING_FILE}（仅处理第 2/3 类）")

    wb = openpyxl.load_workbook(db_path)
    ws = get_actor_db_sheet(wb)

    # 收集待删行号
    del_rows: set[int] = set()
    desc_found: list[tuple[int, str]] = []
    non_acting_found: list[tuple[int, str, int]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        jp = str(row[0] or "").strip()
        tid_val = str(row[5] or "").strip() if len(row) > 5 else ""
        if not jp:
            continue
        if tid_val.isdigit() and int(tid_val) in non_acting_ids:
            del_rows.add(row_idx)
            non_acting_found.append((row_idx, jp, int(tid_val)))
        elif jp in _NON_AV_NAMES:
            del_rows.add(row_idx)
            non_acting_found.append((row_idx, jp, int(tid_val) if tid_val.isdigit() else 0))
        elif not tid_val.isdigit() and jp in _DESC_ROWS:
            del_rows.add(row_idx)
            desc_found.append((row_idx, jp))

    print(f"第 1 类（导演/幕后）待删: {len(non_acting_found)}")
    print(f"第 2 类（描述词）待删: {len(desc_found)}")
    print(f"合计: {len(del_rows)} 行")

    if not args.apply:
        print("\n（预览模式，加 --apply 执行删除）")
        wb.close()
        return 0

    # /tmp 保留原始库副本（工作区安全网，不进仓库）
    shutil.copy(db_path, Path("/tmp/actor_db_before_clean_nonactors.xlsx"))

    # 待删行合并为连续区间，一次 delete_rows(start, count) 批量删除
    # （逐行 delete_rows 每次 O(n) 移动，几百行会超时且中途 kill 会写坏文件）
    sorted_rows = sorted(del_rows, reverse=True)
    intervals: list[tuple[int, int]] = []  # (start, count)
    for row_idx in sorted_rows:
        if intervals and intervals[-1][0] - 1 == row_idx:
            intervals[-1] = (row_idx, intervals[-1][1] + 1)
        else:
            intervals.append((row_idx, 1))
    for start, count in intervals:
        ws.delete_rows(start, count)

    # openpyxl delete_rows 会在末尾保留带格式/超链接的空行，导致 max_row 延伸出空行，
    # 需二次清理（从后往前删 jp 为空的物理行）
    while ws.max_row >= 2:
        last = ws.cell(row=ws.max_row, column=1).value
        if last is None or str(last).strip() == "":
            ws.delete_rows(ws.max_row, 1)
        else:
            break

    wb.save(db_path)
    wb.close()
    print(f"✅ 已删除 {len(del_rows)} 行")
    return 0


if __name__ == "__main__":
    sys.exit(main())
