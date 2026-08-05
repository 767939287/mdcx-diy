"""清洗 actor_database.xlsx 字段噪声。

针对女优库中混入的多类污染：
1. 日文原名/中文名/繁体名 = 占位符或描述词（素人/人妻/女優情報/复元 等非人名）
2. 中文名/繁体名 = 名字+年龄标注（如「涼子 20歳」）或作品系列名
3. 别名 = 混入作品标题（ラグジュTV / ママ友喰い / VOL.xx 等整段标题）
4. 名字 = 混入系列/站点标签（パコパコママ / FC2 / ロリ主婦 / 1000人斬り / 天然むすめ 等）
5. 名字 = 混入年份（（2015）/【2016年】等）
6. 名字 = 纯占位符（FC2 / 素人 / 抜群なアイドル店員 等非人名）
7. 别名 = 残留人妻/熟女/着エロ 等类型标签
8. 别名 = 悬空斜杠/残括号（ただえりさ /、真東愛 / Mahigashi Ai））
9. 简介 = 1-2 字符碎片 → 置空

仅对发生污染的行做字段级修正，不删除任何行。输出统计报告。
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "resources" / "userdata" / "actor_database.xlsx"

PLACEHOLDER_TERMS = [
    "素人", "複数", "复素", "女優情報", "美人な人妻", "人妻", "熟女",
    "管理者様", "復元してください", "复元してください", "凛女",
    "アイドル店員", "ツインテール", "元Fカップ", "グラドル", "愛奴",
    "バド部", "ギャル２人組", "ギャル2人組", "編集", "编集", "奥さん",
    "元Fカップグラドル", "FC2", "生意気ツインテール18ちゃん", "生意気ツインテール",
]

PLACEHOLDER_RE = re.compile("|".join(map(re.escape, PLACEHOLDER_TERMS)))

AGE_RE = re.compile(r"\d+[歳才岁歲]")

# 系列/站点标签：括号内已知非人名，剥离
SERIES_TAG_RE = re.compile(
    r"パコパコママ|エッチな0930|エッチな4610|天然むすめ|ラグジュTV|FC2|1000人斬り|人妻斬り|ロリ主婦|人妻DX|カリビアンコムプレミアム|マドンナ|グラドル|ソープ嬢|トリプルエックス|ニューハーフ|クリスタル|GirlsDelta|無垢|カリビアン|ガチん娘|ムラムラ|DMM素人動画|FC2ライブ|元Fカップ",
    re.IGNORECASE,
)

# 年份标签：括号内纯年份，剥离
YEAR_TAG_RE = re.compile(r"【?\d{4}年?】?")

SERIES_TITLE_RE = re.compile(
    r"ラグジュTV|ママ友喰い|VOL\.?\s*\d+|人妻湯恋旅行|おばさんを酔わせて|相席居酒屋|ナマ交尾|デカクリトリス|イン〇タ|チ〇ポ|オナホ扱い|社交辞令SEX|タダマン|立ちんぼ|美魔女|閉経|生殖活動|メンエス|回春|四畳半|生ハメ|素人限定|素人初撮り|B級熟女|B級素人|芸能人り|元地方局アナ|エッチな0930|人妻斬り|刹那的背徳旅行|パコパコママ|ロリ主婦|天然むすめ|一本道|1000人斬り|千春",
    re.IGNORECASE,
)

# 别名段含「名字+年龄」标注（如 あいみ 20歳 / 神田さん20歳 / 真優38歳）
AGE_SEG_RE = re.compile(r"\d+[歳才岁歲]")

# 别名段为长日文句子（作品标题特征，真实别名不会这么长）
LONG_SEG_RE = re.compile(r"[\u3040-\u30ff\u4e00-\u9fff\uac00-\ud7af]")

# 纯标签残留（别名里单独的 人妻/着エロ/素人 等）
PURE_TAG_RE = re.compile(r"^(人妻|訳あり人妻|熟女|素人|着エロ|人妻DX|主婦|パート)$")

# 已知真别名（括号内容应移入别名列而非删除）
KNOWN_ALIAS_IN_PAREN = re.compile(r"[（(](はんな|結城あかり|瑞穂このみ|アニー麗)[)）]")

# 从名字剥离系列标签/年份（保留名字主体）
def _strip_series_year(name: str) -> str:
    """剥离名字中的系列标签和年份，如 本田仁美(パコパコママ) -> 本田仁美"""
    s = str(name)
    # 移出已知真别名到返回值
    # 剥离 (xxx) / （xxx） 中命中系列标签或年份的部分
    s = re.sub(
        r"[\(（]([^\(（）]*?)[\)）]",
        lambda m: "" if (SERIES_TAG_RE.search(m.group(1)) or YEAR_TAG_RE.fullmatch(m.group(1))) else m.group(0),
        s,
    )
    # 剥离括号后的孤立年份如（2015）
    s = re.sub(r"[\(（\[]\d{4}年?[\)）\]]", "", s)
    s = re.sub(r"[【\[]\d{4}年?[】\]]", "", s)
    # 剥离末尾/开头的年份字面（如 可愛ゆう（2015）已处理，还有 素人2015 形式）
    s = re.sub(r"\d{4}年?$", "", s)
    # 剥离无括号的 FC2ライブ 后缀（如 ゆいかFC2ライブ -> ゆいか）
    s = re.sub(r"FC2ライブ$", "", s)
    s = re.sub(r"元Fカップグラドル$", "", s)
    return s.strip()


# 标注词黑名单：括号内容命中则剥离（国籍/事务所/类型/品牌/组合/本名/年份等）
ANNOTATION_TERMS = re.compile(
    r"英国|ハンガリー|ベトナム|イングランド|俄罗斯混血|USA|JPN|泰国|美国|韩国"
    r"|JETSTREAM|T-POWERS|HEYZO|FALENO|LINX|Gcolle|KUKI|kira☆kira|RealShodo|GOT刊|Playboy|Fleur|きらきら"
    r"|着エロ|ヌードイメージ|女王様|嫁|同人モデル|仮|TS|登録|シンデレラオーディショングランプリ"
    r"|BAND-MAID|DIALOGUE＋|本名|2代目|第\d+期生|デビュー|ニューハーフ"
)


def _strip_annotation(name: str) -> str:
    """剥离名字/别名中的标注括号（国籍/事务所/类型/品牌等），保留别名/读音括号。

    HIBIKI（女王様） -> HIBIKI；Tanaka Karen (田中可恋) 保留（读音/别名）
    """
    s = str(name).strip()
    # 规范化所有括号类型为半角，便于配对
    s = s.replace("【", "[").replace("】", "]").replace("（", "(").replace("）", ")")

    def _should_strip(inner: str) -> bool:
        inner = inner.strip()
        if not inner:
            return True
        # 纯年份（如 2015 / 2016年）→ 剥离
        if YEAR_TAG_RE.fullmatch(inner):
            return True
        # 已知系列/站点标签 → 剥离（即使含假名，如 ロリ主婦/天然むすめ）
        if SERIES_TAG_RE.search(inner):
            return True
        # 标注词命中（即使含假名，如 シンデレラオーディショングランプリ 选秀名）→ 剥离
        if ANNOTATION_TERMS.search(inner):
            return True
        # 含日文假名/韩文 → 视为读音/别名，保留
        if re.search(r"[\u3040-\u30ff\uac00-\ud7af]", inner):
            return False
        return False

    # 处理 [x] 形式
    s = re.sub(r"\[([^\[\]]*)\]", lambda m: "" if _should_strip(m.group(1)) else m.group(0), s)
    # 处理 (x) 形式
    s = re.sub(r"\(([^()]*)\)", lambda m: "" if _should_strip(m.group(1)) else m.group(0), s)
    # 清理未配对的孤立括号（不影响已正确保留的成对括号）
    if s.count("[") != s.count("]") or s.count("(") != s.count(")"):
        s = re.sub(r"[\[\]\(\)]", "", s)
    # 清理孤立标注字面（如 和栗ゆゆ【登録 -> 和栗ゆゆ登録 -> 和栗ゆゆ）
    s = s.replace("登録", "").replace("女王様", "")
    return s.strip()


def _is_placeholder_name(name: str) -> bool:
    """整个名字都是占位符/描述词才返回 True（如 素人/美人な人妻/抜群なアイドル店員）"""
    if not name:
        return False
    s = str(name).strip()
    # 完整命中已知占位符短语 → 直接 True
    for term in ("生意気ツインテール18ちゃん", "抜群なアイドル店員", "複数の素人娘",
                 "复数の素人娘", "美人な人妻", "元Fカップグラドル", "モデルボディーの女",
                 "地味な眼鏡の巨乳妻", "訳アリ巨乳JD", "定時制ギャル", "ギャルママ柔道家",
                 "生意気ツインテール", "人妻湯恋旅行"):
        if term in s:
            return True
    # 去掉括号标签后看主体是否还是占位符
    body = re.sub(r"[\(（][^\(（）]*?[\)）]", "", s).strip()
    if AGE_RE.search(body):
        return True
    if len(body) <= 4 and PLACEHOLDER_RE.search(body):
        return True
    # 占位符词占主体大部分（含占位符词且其余部分也是占位符词）
    if PLACEHOLDER_RE.search(body) and len(re.sub(r"[\u3040-\u30ff\u4e00-\u9fff·・\s]", "", body)) <= 1:
        return True
    return False


def _is_noise_segment(seg: str) -> bool:
    if not seg:
        return False
    if len(seg) <= 5:
        return False
    if AGE_SEG_RE.search(seg):
        return True
    if SERIES_TITLE_RE.search(seg):
        return True
    # 超长日文/中文句子（>20字符）视为作品标题
    if len(seg) > 20 and LONG_SEG_RE.search(seg):
        return True
    return False


def _clean_aliases(alias: str) -> str:
    parts = [p.strip() for p in str(alias).split(",")]
    clean = []
    for p in parts:
        if not p:
            continue
        if PURE_TAG_RE.match(p):
            continue
        if _is_noise_segment(p):
            continue
        # 段内剥离系列标签括号（如 本田仁美(パコパコママ) -> 本田仁美）
        stripped = _strip_series_year(p)
        # 剥离标注括号（如 みゆき（KUKI）/ 工藤唯(着エロ) / 小泉彩（2003年デビュー））
        stripped = _strip_annotation(stripped)
        # 修复悬空斜杠/残括号（如 ただえりさ / -> ただえりさ；真東愛 / Mahigashi Ai） -> 真東愛 / Mahigashi Ai）
        stripped = _fix_dangling_slash(stripped)
        # 日文段内 '名字 + 描述/作品' 空格污染（如 門脇晶子 禁断中出し契約交尾）
        if " " in stripped and not re.search(r"[A-Za-z]", stripped):
            desc_stripped = _strip_desc_tokens(stripped)
            if desc_stripped and desc_stripped != stripped:
                stripped = desc_stripped
            elif not desc_stripped:
                continue  # 整段都是描述词 → 丢弃
        # 紧凑式描述污染（如 巨乳女子プロレスラー凛叶 / ここな先生 / 素人あいか）
        if not re.search(r"[A-Za-z]", stripped):
            compact_stripped = _strip_compact_desc(stripped)
            if compact_stripped != stripped:
                if not compact_stripped:
                    continue  # 纯描述段 → 丢弃
                stripped = compact_stripped
        clean.append(stripped)
    return ",".join(clean)


# 别名段 = '名字 + 作品/描述' 污染（空格分隔，如 門脇晶子 禁断中出し契約交尾）
DESC_TOKEN_RE = re.compile(
    r"店の女|契約|交尾|プロレスラー|の女|禁断|巨乳|耳かき|バニーコレクション|の妻|ナンパ|愛人|の義母|の生徒|の同僚|の幼馴染|の彼女|顔出し|ギャル|制服|メイド|看護師|店員|先生|会長|の娘|素人|熟女|人妻|ランジェリーナ|世田谷の妻|淫乱|レーベル|在宅ワーカー|家賃滞納|いいなり|温泉旅行|ワリキリ|ワリキリバイト|バイト|湘南の女|発禁|患者|万引き娘|夫から逃げる",
)


def _strip_desc_tokens(seg: str) -> str:
    """剥离段中命中的描述/作品 token，保留名字部分。如 門脇晶子 禁断中出し契約交尾 -> 門脇晶子"""
    parts = [p.strip() for p in seg.split(" ") if p.strip()]
    kept = [p for p in parts if not DESC_TOKEN_RE.search(p)]
    if kept:
        return " ".join(kept)
    return ""  # 全部是描述词（如 素人 患者）→ 空


# 紧凑式描述污染（无空格）：前缀/后缀描述词 + 名字
DESC_PREFIX_RE = re.compile(
    r"^(巨乳女子プロレスラー|巨尻女子プロレスラー|巨乳ヒール女子プロレスラー|女子プロレスラー|素人|S級素人|S級色白美肌の素人|淫乱変態ＪＤ|モデルボディーの女|地味な眼鏡の巨乳妻|ギャルママ柔道家|訳アリ巨乳JD|定時制ギャル|複数の素人|复数の素人|色白美巨乳Gカップ美女|美人な人妻|抜群なアイドル店員|ギャル２人組|ギャル2人組|素人美熟女ナンパ|素人庭園|しろハメ素人|俺の素人-Z-|E★人妻DX|泌尿器科女医|幼稚園先生|メイドカフェ店員|♀\d+メイドカフェ店員|巨乳アパレル店員|可愛すぎるス○バ店員|色白152cmあざと可愛いコスメ店員)"
)
DESC_SUFFIX_RE = re.compile(
    r"(先生|女医|メイドカフェ店員|ス○バ店員|アパレル店員|コスメ店員|店員|幼稚園先生|美人妻|人妻看護婦|妻たち|人妻|プロレスラー)$|^(幼なじみの|アラフィフ|五十路|三十六歳|36歳)"
)
DESC_PURE_RE = re.compile(
    r"^(定時制ギャル|訳アリ巨乳JD|モデルボディーの女|地味な眼鏡の巨乳妻|ギャルママ柔道家|美人な人妻|複数の素人娘|复数の素人娘|抜群なアイドル店員|ギャル２人組|ギャル2人組|四十路人妻|素人美熟女ナンパ|S級素人|S级素人|素人奥様|素人不明|素人多数|素人娘|素人娘达|素人娘達|素人品評会|素人妻|素人人物不明|素人１|素人1|素人患者|素人 患者|素人 万引き娘)$"
)


def _strip_compact_desc(seg: str) -> str:
    """剥离紧凑式描述污染：巨乳女子プロレスラー凛叶 -> 凛叶；ここな先生 -> ここな"""
    if not seg:
        return seg
    if DESC_PURE_RE.match(seg):
        return ""
    s = seg
    m = DESC_PREFIX_RE.match(s)
    if m:
        s = s[m.end():].strip()
    m2 = DESC_SUFFIX_RE.search(s)
    if m2:
        s = s[: m2.start()].strip()
    return s.strip()


def _fix_dangling_slash(seg: str) -> str:
    """修复别名段中的悬空斜杠或残括号：ただえりさ / -> ただえりさ；真東愛 / Mahigashi Ai） -> 真東愛 / Mahigashi Ai"""
    if not seg or "/" not in seg and "／" not in seg:
        return seg
    s = seg
    for slash in ("/", "／"):
        if slash in s:
            lhs, _, rhs = s.partition(slash)
            rhs = rhs.strip()
            # 右侧为空 或 右侧只剩残括号 → 保留左侧
            if not rhs or rhs in {")", "）", "]", "】"}:
                s = lhs.strip()
            else:
                # 右侧带残括号（如 Mahigashi Ai））→ 去掉
                s = f"{lhs.strip()} {slash} {re.sub(r'[）)】\]\s]*$', '', rhs).strip()}"
            break
    return s.strip()


def main() -> int:
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["演员数据库"]

    stat = {"jp_placeholder": 0, "cn_placeholder": 0, "tw_placeholder": 0,
            "alias_title": 0, "name_tag": 0, "name_year": 0, "name_placeholder": 0,
            "alias_tag": 0, "bio_short": 0, "alias_extract": 0, "slash_fix": 0,
            "name_annotation": 0}
    detail: list[tuple[str, int, str, str]] = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        jp = row[0].value
        cn = row[1].value
        tw = row[2].value
        alias = row[3].value
        bio = row[8].value
        name_label = jp or cn or tw or "(无名)"

        # 1) 名字字段占位符/年龄 → 置空（用当前单元格值判断，避免旧值回填）
        for idx in (1, 2, 3):
            cur = row[idx - 1].value
            if cur is not None and _is_placeholder_name(str(cur)):
                row[idx - 1].value = None
                stat[["jp_placeholder", "cn_placeholder", "tw_placeholder"][idx - 1]] += 1
                detail.append((name_label, idx, str(cur)[:50], "(置空)"))

        # 2) 名字混入系列标签/年份 → 剥离保留主体
        for idx in (1, 2, 3):
            cur = row[idx - 1].value
            if cur is not None and (SERIES_TAG_RE.search(str(cur)) or YEAR_TAG_RE.search(str(cur))):
                cleaned = _strip_series_year(str(cur))
                if cleaned and cleaned != str(cur):
                    row[idx - 1].value = cleaned
                    stat["name_tag"] += 1
                    detail.append((name_label, idx, str(cur)[:50], cleaned[:50]))

        # 2b) 名字混入标注（国籍/事务所/类型/品牌等，含无括号的如 京子女王様）→ 剥离标注
        for idx in (1, 2, 3):
            cur = row[idx - 1].value
            if cur is not None:
                cleaned = _strip_annotation(str(cur))
                if cleaned and cleaned != str(cur):
                    row[idx - 1].value = cleaned
                    stat["name_annotation"] += 1
                    detail.append((name_label, idx, str(cur)[:50], cleaned[:50]))

        # 3) 名字含已知真别名（括号内容移入别名列）
        for idx, val in ((1, jp), (2, cn), (3, tw)):
            if val:
                m = KNOWN_ALIAS_IN_PAREN.search(str(val))
                if m:
                    alias_part = m.group(1)
                    cleaned = _strip_series_year(str(val))
                    cleaned = cleaned.replace(f"（{alias_part}）", "").replace(f"({alias_part})", "").strip()
                    if cleaned:
                        row[idx - 1].value = cleaned
                        existing_alias = row[3].value or ""
                        new_alias = f"{existing_alias},{alias_part}".strip(",") if existing_alias else alias_part
                        row[3].value = new_alias
                        stat["alias_extract"] += 1
                        detail.append((name_label, idx, str(val)[:40], f"{cleaned[:20]} +别名{alias_part}"))

        # 4) 别名剔除作品标题 + 纯标签 + 修复悬空斜杠
        if alias:
            # 先记录是否有斜杠异常
            had_slash_issue = False
            for _seg in str(alias).split(","):
                _seg = _seg.strip()
                if "/" not in _seg and "／" not in _seg:
                    continue
                _rhs = _seg.split("/")[-1].split("／")[-1].strip()
                if (not _rhs or _rhs in {")", "）", "]", "】"} or re.search(r"[）)】\]\u3000]$", _rhs)):
                    had_slash_issue = True
                    break
            cleaned = _clean_aliases(str(alias))
            if cleaned != str(alias):
                row[3].value = cleaned
                stat["alias_title"] += 1
                if had_slash_issue:
                    stat["slash_fix"] += 1
                detail.append((name_label, 4, str(alias)[:50], cleaned[:50] or "(置空)"))

        # 5) 纯占位符名字（整行名字全是占位符，用当前值判断）
        for idx in (1, 2, 3):
            cur = row[idx - 1].value
            if cur is not None and _is_placeholder_name(str(cur)) and len(str(cur).strip()) <= 8:
                row[idx - 1].value = None
                stat["name_placeholder"] += 1

        # 6) 简介 1-2 字符碎片 → 置空
        if bio and len(str(bio).strip()) <= 2:
            row[8].value = None
            stat["bio_short"] += 1

    wb.save(DB_PATH)

    print(f"📊 清洗报告 (共处理 {ws.max_row - 1} 行)")
    print(f"  日文原名占位符置空: {stat['jp_placeholder']}")
    print(f"  中文名占位符置空: {stat['cn_placeholder']}")
    print(f"  繁体名占位符置空: {stat['tw_placeholder']}")
    print(f"  名字混入系列标签/年份剥离: {stat['name_tag']}")
    print(f"  名字标注括号剥离: {stat['name_annotation']}")
    print(f"  名字真别名移入别名列: {stat['alias_extract']}")
    print(f"  别名剔除作品标题/标签: {stat['alias_title']}")
    print(f"  别名悬空斜杠/残括号修复: {stat['slash_fix']}")
    print(f"  简介碎片置空: {stat['bio_short']}")
    print()
    print("详情(前 25 条):")
    for name, col, old, new in detail[:25]:
        col_name = {1: "日文原名", 2: "中文名", 3: "繁体名", 4: "别名"}[col]
        print(f"  [{name}] {col_name}: {old} -> {new}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
