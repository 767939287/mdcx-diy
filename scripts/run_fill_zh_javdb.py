"""分批执行 fill_zh_javdb，支持断点续传和定时进度报告。

用法：
  uv run python scripts/run_fill_zh_javdb.py [--batch 200] [--state .fill_zh_javdb_state.json]

断点续传：每批完成后把 offset 写入 state 文件，中断后重跑自动从上次位置继续。
"""

import argparse
import asyncio
import json
import os
import sys
import time
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

STATE_FILE = REPO / ".fill_zh_javdb_state.json"
DB_PATH = REPO / "userdata" / "actor_database.xlsx"
DB_ORIG = REPO / "resources" / "userdata" / "actor_database.xlsx"

BATCH_SIZE = 200
REPORT_INTERVAL_SEC = 30  # 每 30 秒打印一次进度

import openpyxl


def count_pending(offset: int) -> int:
    """统计从 offset 行之后还有多少条需要处理（中文名为空或==日文原名，且含汉字）。"""
    wb = openpyxl.load_workbook(DB_PATH, read_only=True)
    ws = wb["演员数据库"]
    count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
        if row_idx - 2 < offset:
            continue
        jp = str(row[0] or "").strip()
        zh_cn = str(row[1] or "").strip()
        if not jp:
            continue
        if (not zh_cn or zh_cn == jp) and any("\u4e00" <= c <= "\u9fff" for c in jp):
            count += 1
    wb.close()
    return count


def count_diff() -> int:
    """统计当前文件相比原始文件有多少行中文名/繁体名发生了变化。"""
    if not DB_ORIG.exists():
        return 0
    wb_old = openpyxl.load_workbook(DB_ORIG, read_only=True)
    wb_new = openpyxl.load_workbook(DB_PATH, read_only=True)
    ws_old = wb_old["演员数据库"]
    ws_new = wb_new["演员数据库"]
    old_rows = list(ws_old.iter_rows(min_row=2, max_col=3, values_only=True))
    new_rows = list(ws_new.iter_rows(min_row=2, max_col=3, values_only=True))
    diff = 0
    for o, n in zip(old_rows, new_rows):
        if str(o[1] or "") != str(n[1] or "") or str(o[2] or "") != str(n[2] or ""):
            diff += 1
    wb_old.close()
    wb_new.close()
    return diff


def load_state(state_file: Path) -> int:
    """返回上次保存的 offset（已处理到的行号），没有则返回 0。"""
    if state_file.exists():
        try:
            data = json.loads(state_file.read_text())
            return data.get("offset", 0)
        except Exception:
            return 0
    return 0


def save_state(state_file: Path, offset: int, total_updated: int) -> None:
    state_file.write_text(json.dumps({"offset": offset, "updated": total_updated}))


async def run_batch(offset: int, batch_size: int) -> int:
    """执行一批，返回实际处理了多少条（用于判断是否跑完）。"""
    from mdcx.tools.actor_db_tool import run_actor_db_xlsx

    await run_actor_db_xlsx("fill_zh_javdb", limit=batch_size, offset=offset)
    return batch_size


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, default=BATCH_SIZE, help="每批处理条数")
    parser.add_argument("--state", type=str, default=str(STATE_FILE), help="状态文件路径")
    args = parser.parse_args()

    state_file = Path(args.state)

    # 确保使用原始文件开始（如果 state 不存在）
    offset = load_state(state_file)
    if offset == 0:
        print("首次运行，复制原始数据库...")
        os.makedirs(DB_PATH.parent, exist_ok=True)
        import shutil

        shutil.copy2(DB_ORIG, DB_PATH)
        print(f"  已复制 {DB_ORIG} -> {DB_PATH}")

    total_pending = count_pending(offset)
    print(f"=== fill_zh_javdb 分批执行 ===")
    print(f"起始 offset: {offset}")
    print(f"待处理条目: {total_pending}")
    print(f"每批: {args.batch}")
    print()

    start_time = time.time()
    last_report = start_time
    total_processed = 0
    total_updated = count_diff()

    batch_num = 0
    while True:
        batch_num += 1
        batch_start = time.time()

        # 检查还有没有待处理
        remaining = count_pending(offset)
        if remaining == 0:
            print(f"\n=== 全部完成！===")
            break

        actual_batch = min(args.batch, remaining)
        print(f"[批次 {batch_num}] offset={offset}, 本批={actual_batch}, 剩余={remaining}")

        await run_batch(offset, actual_batch)

        offset += actual_batch
        total_processed += actual_batch

        # 统计累计更新数
        total_updated = count_diff()
        batch_time = time.time() - batch_start
        elapsed = time.time() - start_time

        save_state(state_file, offset, total_updated)
        print(
            f"  完成: 本批 {batch_time:.1f}s, 累计处理 {total_processed}, "
            f"累计更新 {total_updated}, 总耗时 {elapsed:.0f}s, state 已保存 (offset={offset})"
        )

        # 定时报告
        now = time.time()
        if now - last_report >= REPORT_INTERVAL_SEC:
            rate = total_processed / elapsed if elapsed > 0 else 0
            eta = (remaining - actual_batch) / rate if rate > 0 else 0
            print(f"  [进度报告] 速率 {rate:.1f} 条/s, 预计剩余 {eta:.0f}s ({eta / 60:.1f}min)")
            last_report = now

    elapsed = time.time() - start_time
    print(f"\n=== 最终结果 ===")
    print(f"总处理: {total_processed} 条")
    print(f"总更新: {total_updated} 条 (中文名或繁体名变化)")
    print(f"总耗时: {elapsed:.0f}s ({elapsed / 60:.1f}min)")
    print(f"成功率: {total_updated * 100 / total_processed:.1f}%" if total_processed else "N/A")

    # 清理 state 文件
    if state_file.exists():
        state_file.unlink()
        print("state 文件已清理")


if __name__ == "__main__":
    asyncio.run(main())
