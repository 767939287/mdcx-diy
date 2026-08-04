#!/usr/bin/env python3
"""
演员数据库静态校验脚本。

仅检查仓库内出厂 `resources/userdata/actor_database.xlsx`，把关本地数据质量。
检查项：
  1. jp 空字段（主键必须非空）         [error]
  2. 同 jp 名重复（大小写不敏感）      [error]
  3. keyword 首尾逗号 / 连续逗号       [error]
  4. keyword 重复词（大小写不敏感）    [error]
  5. zh_cn / zh_tw 空字段              [warning]
  6. tmdbid 重复                       [error]
  7. 出生日期列格式（空或 YYYY[-MM[-DD]]） [error]

发现任一 error 返回码 1；仅 warning 返回码 0。老 7 列文件缺失新增列时跳过对应检查。
"""

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover
    sys.stderr.write("缺少依赖 openpyxl，请先 uv sync\n")
    sys.exit(2)

MAIN_PATH = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = MAIN_PATH / "resources" / "userdata" / "actor_database.xlsx"

BIRTH_DATE_PATTERN = re.compile(r"^\d{4}(-\d{1,2}(-\d{1,2})?)?$")


def _check_jp_empty(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip()
        if not jp:
            errors.append(f"  行{idx}: jp(日文原名) 为空")
    return errors


def _check_jp_duplicate(rows):
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        jp = str(row[0] or "").strip().casefold()
        if not jp:
            continue
        if jp in seen:
            errors.append(f"  行{idx}: jp 与行{seen[jp]} 重复: {row[0]}")
        else:
            seen[jp] = idx
    return errors


def _check_keyword_format(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        kw = str(row[3] or "").strip()
        if not kw:
            continue
        if kw.startswith(",") or kw.endswith(",") or ",," in kw:
            errors.append(f"  行{idx}: keyword 存在首尾/连续逗号: {kw}")
    return errors


def _check_keyword_duplicate(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        kw = str(row[3] or "").strip()
        if not kw:
            continue
        parts = [k.strip() for k in kw.split(",") if k.strip()]
        if len(parts) != len({k.casefold() for k in parts}):
            errors.append(f"  行{idx}: keyword 存在重复词: {kw}")
    return errors


def _check_name_empty(rows):
    warnings = []
    for idx, row in enumerate(rows, 2):
        zh_cn = str(row[1] or "").strip() if len(row) > 1 else ""
        zh_tw = str(row[2] or "").strip() if len(row) > 2 else ""
        jp = str(row[0] or "").strip()
        if jp and not zh_cn:
            warnings.append(f"  行{idx}: zh_cn(中文名) 为空")
        if jp and not zh_tw:
            warnings.append(f"  行{idx}: zh_tw(繁体名) 为空")
    return warnings


def _check_tmdbid_duplicate(rows):
    seen: dict[str, int] = {}
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 5:
            continue
        tmdb = str(row[5] or "").strip()
        if not tmdb or not tmdb.isdigit():
            continue
        if tmdb in seen:
            errors.append(f"  行{idx}: tmdbid 与行{seen[tmdb]} 重复: {tmdb}")
        else:
            seen[tmdb] = idx
    return errors


def _check_birth_date(rows):
    errors = []
    for idx, row in enumerate(rows, 2):
        if len(row) <= 7:
            return []  # 老文件无出生日期列，跳过
        birth = str(row[7] or "").strip()
        if birth and not BIRTH_DATE_PATTERN.match(birth):
            errors.append(f"  行{idx}: 出生日期格式非法(期望 YYYY[-MM[-DD]]): {birth}")
    return errors


def check_xlsx(xlsx: Path) -> int:
    if not xlsx.exists():
        print(f"[check_actor_db] 出厂数据库不存在，跳过: {xlsx}")
        return 0

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    errors: list[str] = []
    warnings: list[str] = []
    for check in (
        _check_jp_empty,
        _check_jp_duplicate,
        _check_keyword_format,
        _check_keyword_duplicate,
        _check_tmdbid_duplicate,
        _check_birth_date,
    ):
        errors.extend(check(rows))
    for check in (_check_name_empty,):
        warnings.extend(check(rows))

    print(f"[check_actor_db] {xlsx.relative_to(MAIN_PATH)} 共 {len(rows)} 行数据")
    if errors:
        print("[check_actor_db] 发现 error 级问题:")
        for item in errors:
            print(item)
    if warnings:
        print("[check_actor_db] 发现 warning 级问题(不阻断):")
        for item in warnings:
            print(item)
    if not errors and not warnings:
        print("[check_actor_db] 校验通过")
    elif not errors:
        print("[check_actor_db] 无 error，仅 warning")
    else:
        print(f"[check_actor_db] 校验失败: {len(errors)} 个 error")
    return 1 if errors else 0


def main() -> int:
    parser = argparse.ArgumentParser(description="演员数据库静态校验")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="目标 xlsx 路径")
    args = parser.parse_args()
    return check_xlsx(args.xlsx)


if __name__ == "__main__":
    sys.exit(main())
